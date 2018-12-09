# install package list
# pip3 install selenium (for windows : pip install ---)
# pip3 install bs4

from selenium import webdriver as wd
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook
import time

# get_site : p 태그 다긁어옴, contents_list에 저장
contents_list = []
def get_site(site):
    driver.get(site)
    #
    time.sleep(2)
    #
    contents = driver.find_elements_by_tag_name('p')
    for content in contents:
        contents_list.append(content.text)

# text_preprocessing : 리스트 내용물 다 텍스트로 밀어버림
def text_preprocessing(list):
    all_text = ""
    #
    for text in list:
        all_text = all_text + text
    #
    return all_text

# text_paragraph : . 기준으로 문단 나눠버림, paragraph_list에 저장
def text_paragraph(text):
    paragraph_list = text.split(".")
    return paragraph_list

# save_excel
def save_excel(list):
    wb = Workbook()
    i = 1
    excel = wb.active
    for content in list:
        excel.cell(row=i, column=1, value=i)
        excel.cell(row=i, column=2, value=content)
        i = i + 1
    wb.save('fin.xlsx')

wiki_url =  ['https://en.wikipedia.org/wiki/Korean_regional_cuisine',
            'https://en.wikipedia.org/wiki/Korean_royal_court_cuisine',
            'https://en.wikipedia.org/wiki/List_of_Korean_dishes',
            'https://en.wikipedia.org/wiki/Korean_cuisine']

# driver List_of_Korean_dishes
# for windows : wd.Chrome(executable_path="chromedriver.exe")
#
driver = wd.Chrome(executable_path="./chromedriver")

# site get
for i in range(0, len(wiki_url)):
    get_site(wiki_url[i])

# text preprocessing
text = text_preprocessing(contents_list)

# text_paragraph
list = text_paragraph(text)

# load excel and save
save_excel(list)
